from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell import MergedCell
import re
from datetime import datetime

class ScheduleParser:
    def __init__(self, excel_file):
        self.teachers = dict()
        self.wb = load_workbook(filename=excel_file)
        self.ws = self.wb.active
        self.date_column = 1
        self.first_row = 2
        self.last_date_col = None
        self.groups_row = None

    def get_merged_cell_value(self, cell: Cell):
        if isinstance(cell, MergedCell):
            for merged_range in self.ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left_cell = self.ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    return top_left_cell.value, merged_range
        else:
            return cell.value, None
        return None, None

    def parse_date(self, col: str):
        if col:
            date_match = re.search(r"\d{2}\.\d{2}\.\d{4}", col)
            if date_match:
                date_str = date_match.group()
                date_obj = datetime.strptime(date_str, "%d.%m.%Y")
                return date_obj
        return False

    def parse_subject(self, cell: Cell):
        val, merged_range = self.get_merged_cell_value(cell)
        if val:
            pattern = r'([А-ЯЁа-яё]+ [А-ЯЁ]\.[А-ЯЁ]\.)'
            match = re.search(pattern, val, re.MULTILINE)
            if match:
                parts = re.split(pattern, val)
                subject = parts[0].strip()
                teacher = re.sub(r'\s+', ' ', match.group(1))
                return (subject, teacher, merged_range)
        return False

    def parse_schedule(self):
        max_row = self.ws.max_row
        max_col = self.ws.max_column

        iter_rows = self.ws.iter_rows(
            min_row=self.first_row,
            max_row=max_row,
            min_col=self.date_column,
            max_col=max_col,
        )

        for row in iter_rows:
            date_col = row[0].value
            lesson_num = row[1].value

            dt = None

            if date_col is not None:
                dt = self.parse_date(date_col)
                if dt:
                    self.last_date_col = dt
                else:
                    has_group_row_mark = date_col.find('День')
                    if has_group_row_mark > -1:
                        self.groups_row = row

            else:
                dt = self.last_date_col

            if dt:
                for col in row[2:]:
                    result = self.parse_subject(col)
                    if result:
                        subj, teacher, merged_range = result
                        groups = []
                        if merged_range:
                            start_col = merged_range.min_col
                            end_col = merged_range.max_col
                            for group_col in range(start_col, end_col + 1):
                                group_cell = self.ws.cell(row=self.groups_row[0].row, column=group_col)
                                group_value = group_cell.value
                                if group_value:
                                    for g in group_value.split("\n"):
                                        groups.append(g.strip())
                        else:
                            group_cell = self.ws.cell(row=self.groups_row[0].row, column=col.column)
                            group_value = group_cell.value
                            if group_value:
                                for g in group_value.split("\n"):
                                    groups.append(g.strip())

                        pair = {"subj": subj, "groups": groups, "dt": dt, "lesson_num": lesson_num}
                        if teacher in self.teachers:
                            self.teachers[teacher].append(pair)
                        else:
                            self.teachers[teacher] = [pair]

    def get_teachers_schedule(self):
        return self.teachers
